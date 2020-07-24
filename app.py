import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('testfiile.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)
print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_number = cell.value * 0.5
    corrected_number_cell = sheet.cell(row, 4)
    corrected_number_cell.value = corrected_number

values = Reference(sheet, min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2' )
wb.save('newGENfile2.xlsx')